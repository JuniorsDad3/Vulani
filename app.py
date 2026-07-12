"""
app.py — Vulani backend

Runs the whole app: serves the parent-facing page and the admin
page (from the "public" folder), and stores every submitted
application as a row in a real Excel file (applications.xlsx),
which is created automatically the first time you run this.

WHY EXCEL INSTEAD OF A "REAL" DATABASE
---------------------------------------
For a pilot, this is genuinely fine: applications.xlsx IS the
database. You (or your team) can open it directly in Excel any
time to look at it, and the admin page's export button just hands
you that same file. The trade-off, to be upfront about it: this
only works safely with a SINGLE server process at a time (see
"Running this for real" below) — it's not built to handle dozens
of simultaneous submissions the way a proper database would. For
a pilot with a handful of parents testing it, that's not a
practical problem. If this ever gets to real scale (hundreds of
submissions a day), that's the point to migrate to a proper
database — ask me when you're there, it's a small step from here,
not a rewrite.

HOW TO RUN THIS
----------------
1. Install dependencies (once):
     pip install -r requirements.txt

2. Copy .env.example to .env and fill in your real email details
   (see .env.example for exactly what goes where).

3. Run:
     python app.py

4. Open:
     http://localhost:5000/school-registration-app.html   (parents)
     http://localhost:5000/admin.html                      (your team)

Both pages now load from THIS server (port 5000), not from
"python -m http.server 8000" any more — that was only ever a
stand-in for testing the front end before there was a backend.
"""

import os
import re
import random
import smtplib
import threading
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from flask import Flask, request, jsonify, send_file, send_from_directory
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__, static_folder="public", static_url_path="")

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "applications.xlsx")
write_lock = threading.Lock()  # prevents two submissions corrupting the file if they land at the same instant

ACCESS_CODE = os.getenv("ADMIN_ACCESS_CODE", "changeme123")

MAIL_SERVER = os.getenv("MAIL_SERVER", "smtp.gmail.com")
MAIL_PORT = int(os.getenv("MAIL_PORT", 587))
MAIL_USE_TLS = os.getenv("MAIL_USE_TLS", "True") == "True"
MAIL_USERNAME = os.getenv("MAIL_USERNAME")
MAIL_PASSWORD = os.getenv("MAIL_PASSWORD")
ALERT_EMAIL = os.getenv("ALERT_EMAIL", MAIL_USERNAME)

COLUMNS = [
    "Reference", "Submitted", "Parent Name", "Parent ID", "Parent Phone",
    "Parent Email", "Relationship", "Address", "Child Name", "Child DOB",
    "Grade", "Child ID/Birth Cert", "School", "Town", "Province",
    "Area Type", "Phase",
]


def ensure_workbook():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"
        ws.append(COLUMNS)
        wb.save(EXCEL_PATH)


def append_application(row_dict):
    with write_lock:
        ensure_workbook()
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Applications"]
        ws.append([row_dict.get(col, "") for col in COLUMNS])
        wb.save(EXCEL_PATH)


def read_all_applications():
    ensure_workbook()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Applications"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return [dict(zip(COLUMNS, r)) for r in rows if r and r[0]]


def make_reference():
    stamp = datetime.now().strftime("%y%m%d")
    return f"VLN-{stamp}-{random.randint(1000, 9999)}"


def send_email(to_addr, subject, body):
    if not (MAIL_USERNAME and MAIL_PASSWORD and to_addr):
        return  # silently skip if email isn't configured or no address given
    msg = MIMEMultipart()
    msg["From"] = MAIL_USERNAME
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))
    try:
        with smtplib.SMTP(MAIL_SERVER, MAIL_PORT) as server:
            if MAIL_USE_TLS:
                server.starttls()
            server.login(MAIL_USERNAME, MAIL_PASSWORD)
            server.sendmail(MAIL_USERNAME, to_addr, msg.as_string())
    except Exception as e:
        print(f"[email] failed to send to {to_addr}: {e}")


# ---------------------------------------------------------------
# API
# ---------------------------------------------------------------

@app.route("/api/ping")
def ping():
    return jsonify({"status": "ok"})


@app.route("/api/register", methods=["POST"])
def register():
    data = request.get_json(force=True) or {}

    required = ["pName", "pID", "pPhone", "pAddress", "cName", "cDOB", "cGrade", "school"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"ok": False, "error": f"Missing required fields: {', '.join(missing)}"}), 400

    if not re.fullmatch(r"\d{13}", str(data.get("pID", ""))):
        return jsonify({"ok": False, "error": "Parent ID must be exactly 13 digits."}), 400

    ref = make_reference()
    submitted = datetime.now().strftime("%d %b %Y, %H:%M")
    school = data.get("school", {})

    row = {
        "Reference": ref,
        "Submitted": submitted,
        "Parent Name": data.get("pName", ""),
        "Parent ID": data.get("pID", ""),
        "Parent Phone": data.get("pPhone", ""),
        "Parent Email": data.get("pEmail", ""),
        "Relationship": data.get("pRelation", ""),
        "Address": data.get("pAddress", ""),
        "Child Name": data.get("cName", ""),
        "Child DOB": data.get("cDOB", ""),
        "Grade": data.get("cGrade", ""),
        "Child ID/Birth Cert": data.get("cBirthCert", ""),
        "School": school.get("name", ""),
        "Town": school.get("town", ""),
        "Province": school.get("province", ""),
        "Area Type": school.get("area", ""),
        "Phase": school.get("phase", ""),
    }

    append_application(row)

    body = (
        f"Hi {row['Parent Name']},\n\n"
        f"We've received the application for {row['Child Name']} ({row['Grade']}) "
        f"at {row['School']}.\n\n"
        f"Reference number: {ref}\n"
        f"Submitted: {submitted}\n\n"
        f"Keep this reference number safe — the school will be in touch on "
        f"{row['Parent Phone']} once a place is confirmed. No payment is ever "
        f"required at any step.\n\n— Vulani"
    )
    if row["Parent Email"]:
        send_email(row["Parent Email"], f"Vulani application received — {ref}", body)
    if ALERT_EMAIL:
        send_email(ALERT_EMAIL, f"New application: {row['Child Name']} — {ref}",
                    body + f"\n\nParent phone: {row['Parent Phone']}")

    return jsonify({"ok": True, "ref": ref, "submitted": submitted})


@app.route("/api/applications")
def applications():
    if request.args.get("code") != ACCESS_CODE:
        return jsonify({"error": "unauthorized"}), 401
    return jsonify({"applications": read_all_applications()})


@app.route("/api/export")
def export_excel():
    if request.args.get("code") != ACCESS_CODE:
        return "Unauthorized", 401
    ensure_workbook()
    return send_file(
        EXCEL_PATH,
        as_attachment=True,
        download_name=f"vulani-applications-{datetime.now().strftime('%Y-%m-%d')}.xlsx",
    )


# ---------------------------------------------------------------
# Serve the front-end pages
# ---------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory("public", "school-registration-app.html")


if __name__ == "__main__":
    ensure_workbook()
    app.run(host="0.0.0.0", port=5000, debug=True)
